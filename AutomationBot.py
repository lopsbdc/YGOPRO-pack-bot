from openpyxl import Workbook, load_workbook
from playwright.sync_api import sync_playwright
import time

teste = 0
cartinha = 0
wb = load_workbook("Change the name to your excel file.xlsx") #CHANGE HERE!
ws = wb.active
total_linhas = ws.max_row 
linha = total_linhas + 1

with sync_playwright() as p:
    navegador = p.chromium.launch(headless=False)  #Headless = False. The browser must be active for it to work.

    while teste < 11:
        
        context = navegador.new_context()
        pagina1 = context.new_page()
        pagina1.goto("https://ygoprodeck.com/login/")
        pagina1.locator('//*[@id="displayName"]').fill('INSERT YOUR EMAIL HERE') #CHANGE HERE!
        pagina1.locator('//*[@id="firstPassWord"]').fill('INSERT YOUR PASSWORD HERE!') #CHANGE HERE!
        pagina1.locator('xpath=/html/body/main/div/div/div/article/form/div[5]/button').click()
        print('login successfully!')
        time.sleep(2)
        pagina1.goto("https://ygoprodeck.com/pack-sim/")
        pagina1.locator('//*[@id="filter-dq"]').fill('FILTER THE PACK NAME HERE') #CHANGE HERE!
        time.sleep(2)
        pagina1.locator('xpath=//*[@id="filter-sealedDraft"]').select_option("25") #The bot will open the same pack, 25 times in a row
        time.sleep(2)
        pagina1.locator('xpath=//*[@id="pack-select"]/button').click()
        print('pack selecionado: ' + str(teste))
        
        while cartinha <= 25:
            pagina1.locator('//*[@id="flip"]').click()
            time.sleep(1)
            valor1 = pagina1.locator('xpath=/html/body/main/div/div/div[2]/div[2]/div[8]/div/div[2]/figure/figcaption/span[1]').inner_text()
            valor3 = pagina1.locator('xpath=/html/body/main/div/div/div[2]/div[2]/div[7]/div/div[2]/figure/figcaption/span[1]').inner_text()
            valor2 = pagina1.locator('xpath=/html/body/main/div/div/div[2]/div[2]/div[9]/div/div[2]/figure/figcaption/span[1]').inner_text()
            raridade2 = pagina1.locator('xpath=/html/body/main/div/div/div[2]/div[2]/div[9]/div/div[2]/figure/figcaption/span[2]/span[2]').inner_text()
            raridade1 = pagina1.locator('xpath=/html/body/main/div/div/div[2]/div[2]/div[8]/div/div[2]/figure/figcaption/span[2]/span[2]').inner_text()
            raridade3 = pagina1.locator('xpath=/html/body/main/div/div/div[2]/div[2]/div[7]/div/div[2]/figure/figcaption/span[2]/span[2]').inner_text()
            ws.cell(row=linha, column=2).value = valor1
            ws.cell(row=linha, column=3).value = valor2
            ws.cell(row=linha, column=4).value = valor3

            ws.cell(row=linha, column=7).value = raridade3
            ws.cell(row=linha, column=6).value = raridade2
            ws.cell(row=linha, column=5).value = raridade1
            ws.cell(row=linha, column=1).value = teste
            time.sleep(1)
            cartinha = cartinha + 1
            print('Page ' + str(cartinha) + " was open")
            linha = linha + 1

            if cartinha == 25:
                teste = teste + 1
                cartinha = cartinha + 1
            else:
                pagina1.locator('//*[@id="retry"]').click()
            
        page = teste - 1
        print("Pack " + str(page) + " was open!")
        cartinha = 0
        wb.save("Dados.xlsx")

    print("All packs opened! Waiting for analysis.")
    time.sleep(3000)

